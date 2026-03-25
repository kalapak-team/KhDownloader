import asyncio
import base64
import tempfile
from pathlib import Path

import pdfplumber
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import Workbook

router = APIRouter(prefix="/api/pdf2excel", tags=["pdf2excel"])

MAX_SIZE_MB = 50
MAX_SIZE = MAX_SIZE_MB * 1024 * 1024


@router.post("")
async def pdf_to_excel(
    file: UploadFile = File(..., description="PDF file to convert to Excel"),
) -> JSONResponse:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    ext = Path(file.filename).suffix.lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds maximum allowed size of {MAX_SIZE_MB} MB",
        )

    def _convert() -> bytes:
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "input.pdf"
            src.write_bytes(content)
            out = Path(tmp) / "output.xlsx"

            wb = Workbook()
            # Remove the default sheet — we'll add our own
            wb.remove(wb.active)

            sheet_count = 0

            with pdfplumber.open(str(src)) as pdf:
                for page_idx, page in enumerate(pdf.pages, start=1):
                    tables = page.extract_tables()
                    if tables:
                        for tbl_idx, table in enumerate(tables, start=1):
                            sheet_count += 1
                            name = f"Page{page_idx}"
                            if len(tables) > 1:
                                name = f"Page{page_idx}_T{tbl_idx}"
                            # Sheet name max 31 chars
                            ws = wb.create_sheet(title=name[:31])
                            for row in table:
                                ws.append(
                                    [cell if cell is not None else "" for cell in row]
                                )
                    else:
                        # No tables detected — extract all text as single-column
                        text = page.extract_text()
                        if text and text.strip():
                            sheet_count += 1
                            ws = wb.create_sheet(title=f"Page{page_idx}"[:31])
                            for line in text.split("\n"):
                                ws.append([line])

            if sheet_count == 0:
                raise RuntimeError(
                    "No extractable data found in the PDF"
                )

            wb.save(str(out))
            return out.read_bytes()

    try:
        xlsx_data = await asyncio.to_thread(_convert)
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {exc}") from exc

    stem = Path(file.filename).stem
    safe_stem = (
        "".join(c if ord(c) < 128 else "_" for c in stem).strip("_") or "converted"
    )

    return JSONResponse({
        "filename": f"{safe_stem}.xlsx",
        "data": base64.b64encode(xlsx_data).decode("ascii"),
    })
